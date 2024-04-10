# Determine the Change in Temperature of the common electrode between a heated and probed cell
# The user will input the current and resistance values for the heated and probed cell
# The user will input the time the current is applied
# The user will input the width of the electrode
# The user will input the order of the filament
# The user will select the material of the electrode
# The user can save the output to a text file
# plot the voltage vs current data from an excel file
# Validate the data from ANSYS
import math
import tkinter as tk
from tkinter import filedialog, simpledialog, BooleanVar, ttk
from tkinter.filedialog import askopenfilename
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import os
import pandas as pd

# Find the change in temperature of the common electrode between a heated and probed cell
def QJH_calc():
    global text_to_file_str
    global selected_current_IH_factor
    global selected_current_IP_factor
    global selected_resistance_RonH_factor
    global selected_resistance_RonP_factor
    # ... rest of the function ...
    # Constants
    # electrode distance in microns
    electrode_distance = 150 * (10**-6) # 150 um to meters
    # Thermal conductivity of copper in W/mC or W/mK
    thermal_conductivity_copper = 385
    # Thermal conductivity of platinum in W/mC or W/mK
    thermal_conductivity_platinum = 69.1
    if Copper_electrode.get() and Platinum_electrode.get():
        tk.messagebox.showerror("Error", "Both Copper and Platinum electrodes cannot be selected at the same time.")
        return
    IH_factor = current_IH_factor[selected_current_IH_factor.get()]
    IP_factor = current_IP_factor[selected_current_IP_factor.get()]
    RonH_factor = resistance_RonH_factor[selected_resistance_RonH_factor.get()]
    RonP_factor = resistance_RonP_factor[selected_resistance_RonP_factor.get()]
    # Calculate the Qjh from the new equation
    # Qjh = Q_h + Q_p
    # Get inputs for the heated probe current and the Ron value
    I_H = (float(I_H_entry.get()) * IH_factor) 
    Ron_H = float(Ron_H_entry.get()) * RonH_factor
    # Get inputs from probing cell
    I_P = (float(I_P_entry.get()) * IP_factor) 
    Ron_P = float(Ron_P_entry.get()) * RonP_factor
    # Get time in seconds
    time = float(time_entry.get())
    # Calculate the Q_h
    Q_h = (I_H**2) * Ron_H * time
    # Calculate the Q_p
    Q_p = (I_P**2) * Ron_P * time
    # Calculate the DeltaQ
    DeltaQ = Q_h + Q_p
    # Get electrode width
    width = (float(width_entry.get()) * 10**-6) # convert width to meters
    # Get the order of the filament
    filament_order = float(filament_order_entry.get())
    distance = (width + electrode_distance) * filament_order
    # Choose K from boolean variables based on the user selection
    if Platinum_electrode.get():
        K = thermal_conductivity_platinum
        Across = (width) * (50 * 10**-9) # use meters
        eletrode_material = "Platinum"

    if Copper_electrode.get():
        K = thermal_conductivity_copper
        Across = (width) * (150 * 10**-9) # use meters
        eletrode_material = "Copper"
    # Calculate the DeltaT
    DeltaT = (DeltaQ * distance) / (K * Across * time)
    # round the values to 2 decimal places
    Q_h = round(Q_h, 5)
    Q_p = round(Q_p, 5)
    DeltaQ = round(DeltaQ, 5)
    DeltaT = round(DeltaT, 2)
    input_text_str = (
        "I_H: " + str(I_H) + "\n" +
        "Ron_H: " + str(Ron_H) + "\n" +
        "I_P: " + str(I_P) + "\n" +
        "Ron_P: " + str(Ron_P) + "\n" +
        "Time: " + str(time) + "\n" +  
        "Width: " + str(width) + "\n" +
        "Filament Order: " + str(filament_order) + "\n"
    )
    # Concatenate all the strings
    output_text_str = (
        "The selected material is: " + str(eletrode_material) + "\n" +
        "The difference between the Heated" "\n" +
        "and Probed cell is: " + str(DeltaT) + " Centigrade" + "\n" 
        )
    # combine the input and output strings
    text_to_file_str = (input_text_str + "\n" + output_text_str)
    # Create a text widget to display the output
    output_text = tk.Text(root, width=40, height=3, wrap=tk.WORD, bg='gray', fg='black')
    output_text.grid(row=10, column=0)
    # Insert the result into the text widget
    output_text.delete(1.0, tk.END)
    output_text.insert(tk.END, output_text_str)

# Save the output to a text file
def save_to_file():
    global text_to_file_str
    # Get the filename to save the text to
    file_name = filedialog.asksaveasfilename(filetypes=[("Text files", "*.txt")])
    if file_name:
        # Ensure the file has a .txt extension
        if not file_name.endswith('.txt'):
            file_name += '.txt'
        with open(file_name, "w") as file:
            file.write(text_to_file_str)

# Plot the voltage vs current data from an excel file
def Graph_from_Excel():
    file_name = askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    wb = openpyxl.load_workbook(file_name)
    ws = wb['Sheet1']

    column_names = [cell.value for cell in ws[1]]
        
    # Find the indices of the columns you're interested in
    try:
        voltage_column_index = column_names.index('AV') + 1  # +1 because openpyxl uses 1-based indexing
        current_column_index = column_names.index('AI') + 1
        time_column_index = column_names.index('Time') + 1
    except ValueError:
        tk.messagebox.showerror("Error", "Excel file does not contain 'AV' and 'AI' columns")
        return
        
    # Grab Voltage from column AV
    voltage = []
    for i in range(2, ws.max_row + 1):
        voltage.append(float(ws.cell(row=i, column=voltage_column_index).value))

    # Grab Current from Column AI
    current = []
    for i in range(2, ws.max_row + 1):
        current.append(float(ws.cell(row=i, column=current_column_index).value))

     # Grab Time from Column Time
    time = []
    for i in range(2, ws.max_row + 1):
        time.append(float(ws.cell(row=i, column=time_column_index).value))

    # The reset voltage is the voltage where the current suddenly drops in value
     # find the reset voltage
    for i in range(1, len(current)):
          if current[i] > current[0]:
            reset_voltage = - (voltage[i + 1])
            break
    # Covert microamps to nanoamps for graph
    current = [i * 10**3 for i in current]
    # Plot the voltage vs current data
    fig, ax = plt.subplots()
    fig, ax = plt.subplots()
    fig.patch.set_facecolor('gray')
    ax.set_facecolor('gray')
    ax.spines['bottom'].set_color('darkblue')
    ax.spines['top'].set_color('darkblue') 
    ax.spines['right'].set_color('darkblue')
    ax.spines['left'].set_color('darkblue')
    ax.xaxis.label.set_color('darkblue')
    ax.yaxis.label.set_color('darkblue')
    ax.tick_params(axis='x', colors='darkblue')
    ax.tick_params(axis='y', colors='darkblue')
    ax.plot(voltage, current, 'r')
    ax.set(xlabel='Voltage (V)', ylabel='Current (nA)', title='Voltage vs Current')
    ax.grid()
    # Create the canvas to display the graph
    canvas = FigureCanvasTkAgg(fig, master=root)
    canvas.draw()
    canvas.get_tk_widget().grid(row=12, column=0)
    # Create a text widget to display the output
    output_text = tk.Text(root, width=35, height=4, wrap=tk.WORD, bg='gray', fg='black')
    output_text.grid(row=11, column=0)
    output_text.delete(1.0, tk.END)
    reset_voltage = round(reset_voltage, 2)
    output_text.insert(tk.END, "File: " + file_name + "\n" +
                       "The reset voltage is: " + str(reset_voltage) +" Volts" "\n")
    
# Validate the data from ANSYS
def Validate_from_ANSYS():
    # open system dialog to select text file
    file_path = filedialog.askopenfilename(title="Select Text File", filetypes=[("Text Files", "*.txt")])
    # open the file
    times = []  # List to store all time values
    maximums = []  # List to store all maximum values
    max_temp = float('-inf')  # Initialize max value to negative infinity
    max_time = None  # Initialize max time to None
    with open(file_path, 'r') as file:
        for line in file:
            # Split the line into columns based on whitespace
            columns = line.strip().split()

            # Extract values from columns
            time = columns[0]
            maximum = float(columns[2])  # Assuming index 2 is for the maximum column

            # Append time and maximum values to their respective lists
            times.append(time)
            maximums.append(maximum)

            # Update max value and associated time if current maximum is greater
            if maximum > max_temp:
                max_temp = maximum
                max_time = time
    # Compute Simulated Joule Heating
    # Constants
    # room temperature in Celsius
    room_temperature = 22
    # electrode height in meters
    electrode_height = 25 * (10**-9) # 25 nm to meters
    # Thermal conductivity of the copper ion nanofilament in W/mC or W/mK
    thermal_conductivity_copper = 300
    # electrode radius in meters
    electrode_radius = 10 * (10**-9) # 10 nm to meters
    # Calculate the Qjh from the new equation
    Q = ((max_temp - room_temperature) * ((math.pi) * (electrode_radius**2)) * float(max_time) * 300) / electrode_height
    if Q < 1e-9:
        unit = "pJ"
        Q *= 1e12
    elif Q < 1e-6:
        unit = "nJ"
        Q *= 1e9
    elif Q < 1e-3:
        unit = "uJ"
        Q *= 1e6
    elif Q < 1:
        unit = "mJ"
        Q *= 1e3
    else:
        unit = "J"
    # Resample the time values to standardized the graph display to intervals being every .25 seconds
    # Create a DataFrame from your data
    df = pd.DataFrame({
        'Time': times,
        'Temperature': maximums
    })
    # Convert the 'Time' column to datetime
    df['Time'] = pd.to_datetime(df['Time'], unit='s')
    # Set 'Time' as the index
    df = df.set_index('Time')

    # Resample the data to have a timestep of 0.25 seconds
    df_resampled = df.resample('250ms').mean()

    # Display the results
    output_text = tk.Text(root, width=45, height=5, wrap=tk.WORD, bg='gray', fg='black')
    output_text.grid(row=11, column=1)
    output_text.delete(1.0, tk.END)
    output_text.insert(tk.END, str(file_path) + "\n" +
                       "Maximum Temperature: " + str(max_temp) + "\n" +
                       "Time of Maximum Temperature: " + str(max_time) + "\n" +
                       "Simulated Joule Heating: " + str(round(Q, 2)) + " " + unit + "\n")
    # create a canvas to display the graph
    fig, ax = plt.subplots()
    fig.patch.set_facecolor('gray')
    ax.set_facecolor('gray')
    ax.spines['bottom'].set_color('darkblue')
    ax.spines['top'].set_color('darkblue')
    ax.spines['right'].set_color('darkblue')
    ax.spines['left'].set_color('darkblue')
    ax.xaxis.label.set_color('darkblue')
    ax.yaxis.label.set_color('darkblue')
    ax.tick_params(axis='x', colors='darkblue')
    ax.tick_params(axis='y', colors='darkblue')
    ax.plot(df_resampled.index, df_resampled['Temperature'], 'r')
    ax.set(xlabel='Time (s)', ylabel='Temperature (C)', title='Temperature vs Time')
    ax.grid()
    canvas = FigureCanvasTkAgg(fig, master=root)
    canvas.draw()
    canvas.get_tk_widget().grid(row=12, column=1)

# Main function, create the GUI, buttons, and labels
if __name__ == "__main__":
    root = tk.Tk()
    root.title("GUI")
    root.configure(bg='gray')

    # Create a combobox for the current and resistance factors
    global selected_current_IH_factor
    current_IH_factor = {'mA': 1e-3 , 'uA': 1e-6, 'nA': 1e-9}
    selected_current_IH_factor = tk.StringVar()
    current_IH_dropdown = ttk.Combobox(root, textvariable=selected_current_IH_factor)
    current_IH_dropdown.grid(row=0, column=2)
    global selected_current_IP_factor
    current_IP_factor = {'mA': 1e-3 , 'uA': 1e-6, 'nA': 1e-9}
    selected_current_IP_factor = tk.StringVar()
    current_IP_dropdown = ttk.Combobox(root, textvariable=selected_current_IP_factor)
    current_IP_dropdown.grid(row=2, column=2)
    global selected_resistance_RonH_factor
    resistance_RonH_factor = {'mOhm': 1e-3, 'Ohm': 1, 'kOhm': 1e3, 'MOhm': 1e6}
    selected_resistance_RonH_factor = tk.StringVar()
    resistance_RonH_dropdown = ttk.Combobox(root, textvariable=selected_resistance_RonH_factor)
    resistance_RonH_dropdown.grid(row=1, column=2)
    global selected_resistance_RonP_factor
    resistance_RonP_factor = {'mOhm': 1e-3, 'Ohm': 1, 'kOhm': 1e3, 'MOhm': 1e6}
    selected_resistance_RonP_factor = tk.StringVar()
    resistance_RonP_dropdown = ttk.Combobox(root, textvariable=selected_resistance_RonP_factor)
    resistance_RonP_dropdown.grid(row=3, column=2)

    # Set the values for the comboboxes
    current_IH_dropdown['values'] = tuple(current_IH_factor.keys())
    current_IH_dropdown.current(1)
    current_IP_dropdown['values'] = tuple(current_IP_factor.keys())
    current_IP_dropdown.current(1)
    resistance_RonH_dropdown['values'] = tuple(resistance_RonH_factor.keys())
    resistance_RonH_dropdown.current(1)
    resistance_RonP_dropdown['values'] = tuple(resistance_RonP_factor.keys())
    resistance_RonP_dropdown.current(1)
   
    # Create a menubar
    menubar = tk.Menu(root)
    root.config(menu=menubar)

    # Create a file menu
    file_menu = tk.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="File", menu=file_menu)
    # Add a command to plot the voltage vs current data
    file_menu.add_command(label="Plot From Excel", command=Graph_from_Excel)
    # Add a command to save the output to a file
    file_menu.add_command(label="Save Output", command=save_to_file)
    # Add a command to validate the data from ANSYS
    file_menu.add_command(label="Validate from ANSYS", command=Validate_from_ANSYS)

    # Create label and entry for the I_H value
    I_H_label = tk.Label(root, text="I_H", bg='gray', fg='SpringGreen2')
    I_H_label.grid(row=0, column=0)
    I_H_entry = tk.Entry(root)
    I_H_entry.grid(row=0, column=1)

    # Create label and entry for the Ron_H value
    Ron_H_label = tk.Label(root, text="Ron_H ", bg='gray', fg='firebrick1')
    Ron_H_label.grid(row=1, column=0)
    Ron_H_entry = tk.Entry(root)
    Ron_H_entry.grid(row=1, column=1)

    # Create label and entry for the I_P value
    I_P_label = tk.Label(root, text="I_P", bg='gray', fg='SpringGreen2')
    I_P_label.grid(row=2, column=0)
    I_P_entry = tk.Entry(root)
    I_P_entry.grid(row=2, column=1)

    # Create label and entry for the Ron_P value
    Ron_P_label = tk.Label(root, text="Ron_P", bg='gray', fg='firebrick1')
    Ron_P_label.grid(row=3, column=0)
    Ron_P_entry = tk.Entry(root)
    Ron_P_entry.grid(row=3, column=1)

    # Create label and entry for the time value
    time_label = tk.Label(root, text="Time (seconds)", bg='gray', fg='black')
    time_label.grid(row=4, column=0)
    time_entry = tk.Entry(root)
    time_entry.grid(row=4, column=1)

    # Create label and entry for the width value
    width_label = tk.Label(root, text="Width (microns)", bg='gray', fg='IndianRed1')
    width_label.grid(row=5, column=0)
    width_entry = tk.Entry(root)
    width_entry.grid(row=5, column=1)

    # Create label and entry for the filament order value
    filament_order_label = tk.Label(root, text="Filament Order", bg='gray', fg='black')
    filament_order_label.grid(row=6, column=0)
    filament_order_entry = tk.Entry(root)
    filament_order_entry.grid(row=6, column=1)

    # Create a BooleanVar to hold the state of the checkbox
    Platinum_electrode = BooleanVar()
    Copper_electrode = BooleanVar()
    
    # Create a label and checkbutton for the Platinum electrode
    Platinum_electrode_label = tk.Label(root, text="Platinum Electrode", bg='gray', fg='cyan2')
    Platinum_electrode_label.grid(row=7, column=0)
    Platinum_electrode_checkbutton = tk.Checkbutton(root, variable=Platinum_electrode, bg='gray', fg='blue')
    Platinum_electrode_checkbutton.grid(row=7, column=1)

    # Create a label and checkbutton for the Copper electrode
    Copper_electrode_label = tk.Label(root, text="Copper Electrode", bg='gray', fg='DarkOrange2')
    Copper_electrode_label.grid(row=8, column=0)
    Copper_electrode_checkbutton = tk.Checkbutton(root, variable=Copper_electrode, bg='gray', fg='blue')
    Copper_electrode_checkbutton.grid(row=8, column=1)

    # Create a button to calculate the DeltaT
    calculate_button = tk.Button(root, text="Calculate DeltaT", command=QJH_calc, bg='gray', fg='cyan2')
    calculate_button.grid(row=9, column=0)

    root.protocol("WM_DELETE_WINDOW", root.quit)
    root.mainloop()