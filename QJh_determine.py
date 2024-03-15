# Select an excel file with the voltage and current data
# Calculate the QJh value for the user defined Ron value
# Calculate the temperature of the copper, platinum, and silicon dioxide
# Plot the voltage vs current data
import math
import tkinter as tk
from tkinter import filedialog, simpledialog, BooleanVar
from tkinter.filedialog import askopenfilename
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import openpyxl
import os
import pandas as pd


def open_file():
    global file_name
    file_name = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
    if not file_name:
        return


def QJH_calc():
    global output_text
    global canvas
    global file_name
    # Constants
    # Density of copper in g/cm^3
    copper_density = 8.96
    # Specific heat of copper in J/gC        
    specific_heat_copper = 0.385
    # specific heat capacity of platinum in J/gC
    specific_heat_platinum = 0.134
    # density of platinum in g/cm^3
    platinum_density = 21.45
    # specific heat capacity of Silcon Dioxide in J/gC
    specific_heat_silicon_dioxide = 0.703
    # density of silicon dioxide in g/cm^3
    silicon_dioxide_density = 8.2
    if use_input_voltage.get():
        reset_voltage = float(voltage_entry.get())
    else:
        file_name = askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        wb = openpyxl.load_workbook(file_name)
        ws = wb['Sheet1']

        column_names = [cell.value for cell in ws[1]]
        
        # Find the indices of the columns you're interested in
        try:
            voltage_column_index = column_names.index('AV') + 1  # +1 because openpyxl uses 1-based indexing
            current_column_index = column_names.index('AI') + 1
            time_column_index = column_names.index('Time') + 1
        except ValueError:
            tk.messagebox.showerror("Error", "Excel file does not contain 'AV' and 'AI' columns")
            return
        
        # Grab Voltage from column AV
        voltage = []
        for i in range(2, ws.max_row + 1):
            voltage.append(float(ws.cell(row=i, column=voltage_column_index).value))

        # Grab Current from Column AI
        current = []
        for i in range(2, ws.max_row + 1):
            current.append(float(ws.cell(row=i, column=current_column_index).value))

        # Grab Time from Column Time
        time = []
        for i in range(2, ws.max_row + 1):
            time.append(float(ws.cell(row=i, column=time_column_index).value))

        # The reset voltage is the voltage where the current suddenly drops in value
        # find the reset voltage
        for i in range(1, len(current)):
            if current[i] > current[0]:
                reset_voltage = - (voltage[i + 1])
                break
        # Covert microamps to nanoamps for graph
        current = [i * 10**3 for i in current]

    # Calculate the QJh value for user defined compliance current
    f_dissipation = float(f_dissipation_entry.get())
    RR = float(RR_entry.get())
    Ron = float(Ron_entry.get())
    Qjh = ((reset_voltage**3 ) / (3 * RR * Ron)) * (1 - f_dissipation)
    
    #The radii of the filament in nanometers can be user defined. Currently hard coded as the 10-25 filament
    # dimensions of the truncated cone in nanometers >> centimeters
    r1 = 12.5 * 10**(-7)
    r2 = 5 * 10**(-7)
    h = 25 * 10**(-7)

    # Calculate volume of truncated cone in nm^3
    # R1 is the radius of the larger base        
    # R2 is the radius of the smaller base
    volume_filament = (1/3) * math.pi * h * (r1**2 + r2**2 + r1*r2)

    # Calculate filament mass in grams
    mass_filament = volume_filament * copper_density

    # Copper electrode consists of a single rectangle prism with two pads on the end
    # Copper layer is 150 nm thick
    # Pad is 100um by 100um
    # convert to cm
    #               100um = 100 * 10**(-4) cm, 150nm = 150 * 10**(-7) cm
    copper_pad_volume = ((100 * 10**(-4))**2) * (150 * 10**(-7))
    copper_pad_mass = copper_pad_volume * copper_density
    # Calculate mass of copper electrode in grams
    # 800um  >> 0.08 cm
    volume_copper_electrode = .08 * (10 * 10**(-4)) * (150 * 10**(-7))
    mass_copper_electrode = volume_copper_electrode * copper_density
    total_mass_copper = ( 2 * mass_filament) + (2* copper_pad_mass) + mass_copper_electrode

    # Platinum electrode consists of a single rectangle prism with two pads on the end
    # Platinum layer is 50 nm thick
    # Pad is 100um by 100um
    plat_pad_volume = ((100 * 10**(-4))**2) * (50 * 10**(-7))
    plat_pad_mass = plat_pad_volume * platinum_density
    # Calculate mass of platinum electrode in grams
    volume_platinum_electrode = .08 * (10 * 10**(-4)) * (50 * 10**(-7))
    mass_platinum_electrode = volume_platinum_electrode * platinum_density
    total_mass_platinum = mass_platinum_electrode + (2 * plat_pad_mass)

    # silicon dioxide layer is 25 nm thick
    # silicon dioxide is a single sheet 
    # Calculate mass of silicon dioxide in grams
    # Oxide is 1030um by 750um
    volume_silicon_dioxide = (1030 * 10**(-4)) * (750 * 10**(-4)) * (25 * 10**(-7))
    mass_silicon_dioxide = volume_silicon_dioxide * silicon_dioxide_density

    # Calculate the temperature of the copper, platinum, and silicon dioxide
    Qst = Qjh * .1
    temp = Qjh / ((total_mass_copper * specific_heat_copper) + (total_mass_platinum * specific_heat_platinum) + (mass_silicon_dioxide * specific_heat_silicon_dioxide))
    tempFil = Qst / (mass_filament * specific_heat_copper)


    # Clear the text widget
    output_text.delete(1.0, tk.END)

    # Clear the plot widget
    plt.clf()
    reset_voltage = round(reset_voltage, 2)

    # Display the output in the text widget
    # display the file name
    # Concatenate all the strings
    if 'file_name' in globals():
        file_name_str =  "File: " + os.path.basename(file_name) 
    else:
        file_name_str = "No file selected"
    output_text_str = (
        file_name_str + "\n"
        "The fraction of heat dissipated by convection and thermal radiation is: " + str(f_dissipation_entry.get()) + "\n"
        "The ramping rate is: " + str(RR_entry.get()) + " V/s\n"
        "The reset voltage is: " + str(reset_voltage) + " volts\n"
        "Qjh is: " + str(Qjh) + " micro Joules\n"
        "The Ron value is: " + str(Ron_entry.get()) + " Ohms\n"
        "Accounting for heat removed by convection and thermal radiation\n"
        "the average temperature change of the device is " + str(temp) + " degrees Celsius\n"
        "The temperature change of the filament is " + str(tempFil) + " degrees Celsius\n"
        "Qst is: " + str(Qst) + " micro Joules\n"
        "mass of the filament is: " + str(mass_filament) + " grams\n"
        )
    # Insert the result into the text widget
    output_text.delete(1.0, tk.END)
    output_text.insert(tk.END, output_text_str)

    if not use_input_voltage.get():
        # Create a plot of the voltage vs current
        # Plot the data
        fig, ax = plt.subplots()
        ax.plot(voltage, current)
        ax.set(xlabel='Voltage (V)', ylabel='Current (nA)', title='Voltage vs Current')
        # Create a canvas and add it to the window
        canvas = FigureCanvasTkAgg(fig, master=root)
        canvas.draw()
        canvas.get_tk_widget().grid(row=6, column=0, columnspan=2)
    else:
        if 'canvas' in globals():
            canvas.get_tk_widget().destroy()
    


if __name__ == "__main__":
    root = tk.Tk()
    root.title("GUI")
    
    root.configure(bg='gray14')
    
    # Create label and entry widgets for user input
    RR_label = tk.Label(root, text="Enter the Ramping Rate in V/s", bg='gray14', fg='SpringGreen2')
    RR_label.grid(row=0, column=0)
    RR_entry = tk.Entry(root, width=50)
    RR_entry.grid(row=0, column=1)
    
    # Create a BooleanVar to hold the state of the checkbox
    use_input_voltage = BooleanVar()
    
    # Create a checkbutton to toggle between user input voltage and Excel voltage
    checkbutton = tk.Checkbutton(root, text="Use input voltage", variable=use_input_voltage, bg='gray14', fg='SpringGreen2')
    checkbutton.grid(row=3, column=0)
    
    # Create an entry for the user to input the voltage
    voltage_entry = tk.Entry(root, width=50)
    voltage_entry.grid(row=3, column=1)
    
    # Create a label and entry for the Ron value
    Ron_label = tk.Label(root, text="Enter the Ron value in Ohms", bg='gray14', fg='SpringGreen2')
    Ron_label.grid(row=1, column=0)
    Ron_entry = tk.Entry(root, width=50)
    Ron_entry.grid(row=1, column=1)

    # Create a label and entry for the fraction of heat dissipated
    f_dissipation_label = tk.Label(root, text="Enter the fraction of heat dissipated as a decimal", bg='gray14', fg='SpringGreen2')
    f_dissipation_label.grid(row=2, column=0)
    f_dissipation_entry = tk.Entry(root, width=50)
    f_dissipation_entry.grid(row=2, column=1)

    # Create a calculate button and pack it
    calculate_button = tk.Button(root, text="Calculate", command=QJH_calc)
    calculate_button.grid(row=4, column=1, sticky = 'w')

    # Create a text widget to display the output
    output_text = tk.Text(root, height=15, width=150, fg='SpringGreen2', bg='gray14')
    output_text.grid(row=5, column=0, columnspan=2, )
    
    # Create a canvas to display the plot
    #canvas = tk.Canvas(root)
    #canvas.grid(row=5, column=0, columnspan=2)
    
    root.protocol("WM_DELETE_WINDOW", root.quit)
    root.mainloop()